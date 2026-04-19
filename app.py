import re
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Excel Sheets Merger", layout="wide", page_icon="📊")

st.title("📊 Excel Sheets Merger")
st.markdown("Upload an Excel file — all sheets will be merged into one.")


def find_header_row(df):
    for i in range(len(df)):
        vals = [str(v) for v in df.iloc[i].tolist()]
        if any("رقم" in v or "المبلغ" in v or "الاسم" in v for v in vals):
            return i
    return None


def find_label_value(df, label, search_rows=15):
    """Find row containing `label` in any column, return value immediately to its left."""
    for i in range(min(search_rows, len(df))):
        vals = [str(v) for v in df.iloc[i].tolist()]
        for j, v in enumerate(vals):
            if v.strip() == label and j > 0:
                left = vals[j - 1]
                if left and left != "nan":
                    return left.strip()
    return ""


def extract_data(df, sheet_name):
    header_idx = find_header_row(df)
    if header_idx is None:
        return None, None

    headers = df.iloc[header_idx].tolist()

    # Identify data columns (skip all-NaN leading columns)
    data_cols = [i for i, h in enumerate(headers) if str(h) not in ("nan", "")]
    if not data_cols:
        return None, None

    col_names = {i: str(headers[i]) for i in data_cols}

    data = df.iloc[header_idx + 1:].copy()
    data = data[data_cols].rename(columns=col_names)

    # Drop rows where ALL data columns are NaN (empty rows + total rows)
    data = data.dropna(how="all")

    # Drop the total row: row where the last column (row-number col) is NaN
    row_num_col = col_names[data_cols[-1]]
    data = data[data[row_num_col].notna()]

    # Pull metadata from the top of the sheet
    data["الادارة"] = find_label_value(df, "الادارة")
    jadwal = find_label_value(df, "صنف الموظفين")  # e.g. "جدول 04"
    data["الجدول"] = jadwal
    m = re.search(r"\d+", jadwal) if jadwal else None
    data["رقم الجدول"] = m.group(0) if m else ""
    data["طريقة الدفع"] = find_label_value(df, "طريقة الدفع")
    data["الورقة"] = sheet_name
    return data, col_names


uploaded_file = st.file_uploader("Choose an Excel file (.xlsx / .xls)", type=["xlsx", "xls"])

if uploaded_file is not None:
    xl = pd.ExcelFile(uploaded_file)
    sheet_names = xl.sheet_names

    st.info(f"Found **{len(sheet_names)}** sheets")

    col1, col2 = st.columns([3, 1])
    with col1:
        selected = st.multiselect(
            "Sheets to merge (all selected by default)",
            options=sheet_names,
            default=sheet_names,
        )
    with col2:
        renumber = st.checkbox("Renumber rows (1, 2, 3…)", value=True)
        add_sheet_col = st.checkbox("Keep source sheet column", value=True)

    if selected and st.button("Merge", type="primary"):
        all_frames = []
        skipped = []
        col_names_ref = None

        for name in selected:
            df_raw = pd.read_excel(xl, sheet_name=name, header=None)
            frame, col_names = extract_data(df_raw, name)
            if frame is None:
                skipped.append(name)
                continue
            if col_names_ref is None:
                col_names_ref = col_names
            all_frames.append(frame)

        if skipped:
            st.warning(f"Skipped sheets (no recognizable header): {', '.join(skipped)}")

        if not all_frames:
            st.error("No data could be extracted.")
            st.stop()

        combined = pd.concat(all_frames, ignore_index=True)

        if not add_sheet_col:
            combined = combined.drop(columns=["الورقة"], errors="ignore")

        # Renumber the row-number column
        if renumber and col_names_ref:
            row_num_col = str(list(col_names_ref.values())[-1])
            if row_num_col in combined.columns:
                combined[row_num_col] = range(1, len(combined) + 1)

        st.success(f"Merged **{len(combined)}** rows from **{len(all_frames)}** sheets")

        # Show total amount if present
        amount_col = next((c for c in combined.columns if "المبلغ" in str(c)), None)
        if amount_col:
            total = pd.to_numeric(combined[amount_col], errors="coerce").sum()
            st.metric("Total Amount | المبلغ الإجمالي", f"{total:,.2f}")

        st.dataframe(combined, use_container_width=True, height=420)

        # ── Build output Excel with formatting ──────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "Merged"
        ws.sheet_view.rightToLeft = True  # RTL for Arabic

        header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        total_fill = PatternFill("solid", fgColor="E2EFDA")
        total_font = Font(bold=True, name="Arial", size=11)
        data_align = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        cols = list(combined.columns)

        # Header row
        for c_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
            ws.column_dimensions[get_column_letter(c_idx)].width = 22

        # Data rows
        for r_idx, row in enumerate(combined.itertuples(index=False), start=2):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = data_align
                cell.border = border

        # Total row
        last_data_row = len(combined) + 1
        total_row = last_data_row + 1
        ws.cell(row=total_row, column=1, value="المجموع").font = total_font
        ws.cell(row=total_row, column=1).fill = total_fill
        ws.cell(row=total_row, column=1).border = border

        if amount_col:
            amt_col_idx = cols.index(amount_col) + 1
            amt_letter = get_column_letter(amt_col_idx)
            total_cell = ws.cell(
                row=total_row,
                column=amt_col_idx,
                value=f"=SUM({amt_letter}2:{amt_letter}{last_data_row})",
            )
            total_cell.font = total_font
            total_cell.fill = total_fill
            total_cell.border = border

        ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        original_name = uploaded_file.name.rsplit(".", 1)[0]
        st.download_button(
            label="⬇️ Download Merged File",
            data=output.getvalue(),
            file_name=f"{original_name}_merged.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )
